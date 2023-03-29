from openpyxl import load_workbook, Workbook
from django.shortcuts import render
from .models import Translation
from django.views import View
from django.views.generic.list import ListView
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from .models import Translation, ValueState


class WordsBaseView(View):
    '''базовый класс'''
    model = Translation
    fields = '__all__'
    success_url = reverse_lazy('all')

class WordsListView(WordsBaseView, ListView):
    """Список всех слов"""

class WordsCreateView(WordsBaseView, CreateView):
    """Создание новой пары слов"""

class WordsUpdateView(WordsBaseView, UpdateView):
    """Редактирование пары слов"""

class WordsDeleteView(WordsBaseView, DeleteView):
    """Удаление пары слов"""

def search_word(request):
    user_value  = request.GET.get('search_word')
    last_state = ValueState.objects.last()
    from_bd = Translation.objects.all()
    excel_list, bd_list = work_lists()
    if last_state.excel_state < len(excel_list):
        add_words_from_exel()
        state_value()
    elif last_state.bd_state < len(bd_list):
        add_words_from_DB_in_exel()
        state_value()
    elif last_state.bd_state > len(bd_list):
        new_excel()
        state_value()
    elif last_state.excel_state > len(excel_list):
        Translation.objects.all().delete()
        add_words_from_exel()
        #del_bd()
        state_value()
    else:
        state_value()



    try:
        result_BD = Translation.objects.get(en=user_value)
        if result_BD:
            result = '{0}-{1}'.format(result_BD.en, result_BD.uk)
            return render(request, 'perevod_app/search_word.html', {'result': result, 'from_bd':from_bd})
        else:
            result = 'Слово не знайдено'
        return render(request, 'perevod_app/search_word.html', {'result': result, 'from_bd':from_bd})

    except:
        result = 'Слово не знайдено'
        return render(request, 'perevod_app/search_word.html', {'result': result, 'from_bd':from_bd})

def add_words_from_exel():
    '''сохранение всех записей с exel в БД'''
    wb = load_workbook(filename='translations.xlsx')
    sheet = wb['Sheet1']
    for row in sheet.iter_rows(values_only=True):
        try:
            trans_bd = Translation()
            trans_bd.en=row[0]
            trans_bd.uk=row[1]
            trans_bd.save()
        except:
            continue

def add_words_from_DB_in_exel():
    '''сохранение всех записей с БД в еxel'''
    wb = load_workbook(filename='translations.xlsx')
    sheet = wb['Sheet1']
    trans_bd = Translation.objects.all()
    exel_list =[]
    for row in sheet.iter_rows(values_only=True):
        exel_list.append(row[0])
    for word in trans_bd:
        if word.en not in exel_list:
            sheet.append({'A':word.en, 'B':word.uk})
            wb.save('translations.xlsx')

def work_lists():
    wb = load_workbook(filename='translations.xlsx')
    sheet = wb['Sheet1']
    trans_bd = Translation.objects.all()
    exel_list =[]
    bd_list = []
    for row in sheet.iter_rows(values_only=True):
        exel_list.append(row[0])
    for word in trans_bd:
        bd_list.append(word.en)
    return exel_list, bd_list

def state_value():
    excel_list, bd_list = work_lists()
    state = ValueState()
    state.excel_state = len(excel_list)
    state.bd_state = len(bd_list)
    state.save()


def new_excel():
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Sheet1'
    trans_bd = Translation.objects.all()

    for word in trans_bd:
        sheet.append({'A':word.en, 'B':word.uk})
        wb.save('translations.xlsx')


